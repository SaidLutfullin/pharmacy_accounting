import loguru
from django.db import connection
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, Alignment
from collections import namedtuple
from decimal import Decimal
from src.drugs.models import Shipment, Movement, Distribution, Department


class Report:
    def __init__(self, title, column_titles, values, total_sum_row=None, columns_width=None, is_movement_report=False):
        self.title = title
        self.column_titles = column_titles
        self.values = values
        self.total_sum_row = total_sum_row
        self.columns_width = columns_width
        self.is_movement_report = is_movement_report

    @staticmethod
    def use_by_date_report(date):
        shipments = Shipment.objects.filter(current_amount__gt=0, use_by_date__lt=date)

        title = ['Отчет', f'по лекарственным средствам у которых истёк срок годности на {date.strftime("%d.%m.%Y")} года.']
        column_titles = ['№', 'Наименование', 'Дата получ.', 'Срок годн.', 'Ед.измерения', 'Колич.', 'Цена за ед.', 'Общая сумма']
        amount_of_expired_drugs = 0
        price_of_expired_drugs = Decimal('0')
        values = []
        for index, shipment in enumerate(shipments): 
            value = [index+1,
                     shipment.drug.name,
                     shipment.date_of_comming.strftime("%d.%m.%Y"),
                     shipment.use_by_date.strftime("%d.%m.%Y"),
                     shipment.drug.unit.name,
                     shipment.current_amount,
                     shipment.prise_for_unit,
                     shipment.total_price]
            values.append(value)
            amount_of_expired_drugs += shipment.current_amount
            price_of_expired_drugs += price_of_expired_drugs+shipment.total_price
        total_sum_row = {
            'E': 'Итого:',
            'F': amount_of_expired_drugs,
            'H': price_of_expired_drugs
        }
        columns_width = (3, 36, 11, 11, 13, 9, 11, 17)
        return Report([title], column_titles, [values], [total_sum_row], columns_width)

    @staticmethod
    def presence_report(date):
        shipments = Shipment.objects.filter(current_amount__gt=0).select_related('drug', 'provider',
                                                                                 'producer', 'drug__unit')

        title = ['Отчет', f'наличия лекарственных средств на {date.strftime("%d.%m.%Y")} года.']
        column_titles = ['№', 'Наименование', 'Фирма-поставщик', 'Назв. документа',
                         'Фирма-изготовитель', 'Серия препарата', 'Дата получ.',
                         'Срок годн.', 'Ед.измерения', 'Колич.', 'Цена за ед.', 'Общая сумма']

        values = []
        price_of_present_drugs = 0
        amount_of_present_drugs = Decimal('0')
        for index, shipment in enumerate(shipments): 
            value = [index+1,
                     shipment.drug.name,
                     shipment.provider.name,
                     shipment.document,
                     shipment.producer.name,
                     shipment.serial_number,
                     shipment.date_of_comming.strftime("%d.%m.%Y"),
                     shipment.use_by_date.strftime("%d.%m.%Y"),
                     shipment.drug.unit.name,
                     shipment.current_amount,
                     shipment.prise_for_unit,
                     shipment.total_price]
            values.append(value)
            amount_of_present_drugs += shipment.current_amount
            price_of_present_drugs += price_of_present_drugs+shipment.total_price
        
        total_sum_row = {
            'I': 'Итого:',
            'J': amount_of_present_drugs,
            'L': price_of_present_drugs
        }
        columns_width = (3, 35, 20, 15, 21, 16, 11, 11, 13, 9, 11, 17)
        return Report([title], column_titles, [values], [total_sum_row], columns_width)

    @staticmethod
    def expenditure_report(from_date, to_date, date_of_movement_needed=False):
        movements = Movement.objects.filter(date__gte=from_date, date__lte=to_date)\
                .select_related('shipment', 'shipment__provider', 'shipment__producer',
                                'shipment__drug', 'shipment__drug__unit')
        
        title = ['Отчет', f'израсходованных лекарственных средств с {from_date.strftime("%d.%m.%Y")} по {to_date.strftime("%d.%m.%Y")} года.']
        column_titles = ['№', 'Наименование', 'Фирма-поставщик', 'Назв. документа',
                         'Фирма-изготовитель', 'Серия препарата', 'Дата получ.',
                         'Ед.измерения', 'Колич.', 'Цена за ед.', 'Общая сумма']
        
        values = []
        price_of_expenditure_drugs = 0
        amount_of_expenditure_drugs = Decimal('0')
        for index, movement in enumerate(movements): 
            value = [index+1, movement.shipment.drug.name,
                     movement.shipment.provider.name,
                     movement.shipment.document,
                     movement.shipment.producer.name,
                     movement.shipment.serial_number,
                     movement.shipment.date_of_comming.strftime("%d.%m.%Y"),
                     movement.shipment.drug.unit.name,
                     movement.amount,
                     movement.shipment.prise_for_unit,
                     movement.total_price]
            
            if date_of_movement_needed:
                value.insert(7, movement.date.strftime("%d.%m.%Y"))

            values.append(value)
            amount_of_expenditure_drugs += movement.amount
            price_of_expenditure_drugs += price_of_expenditure_drugs+movement.total_price
        columns_width = [3, 35, 20, 15, 21, 16, 11, 13, 9, 11, 17]
        
        if date_of_movement_needed:
            columns_width.insert(7, 11)
            column_titles.insert(7, 'Дата расх.')
        
            total_sum_row = {
                'I': 'Итого:',
                'J': amount_of_expenditure_drugs,
                'L': price_of_expenditure_drugs}
        else:
            total_sum_row = {'H': 'Итого:',
                             'I': amount_of_expenditure_drugs,
                             'K': price_of_expenditure_drugs}
        return Report([title], column_titles, [values], [total_sum_row], columns_width)

    @staticmethod
    def shipment_report(from_date, to_date):
        shipments = Shipment.objects.filter(date_of_comming__gte=from_date, date_of_comming__lte=to_date)\
            .select_related('drug', 'provider', 'producer', 'drug__unit')

        title = ['Отчет', f'приобретённых лекарственных средств с {from_date.strftime("%d.%m.%Y")} по {to_date.strftime("%d.%m.%Y")} года.']
        
        column_titles = ['№', 'Наименование', 'Фирма-поставщик', 'Назв. документа',
                         'Фирма-изготовитель', 'Серия препарата', 'Дата получ.', 'Срок годн.',
                         'Ед.измерения', 'Колич.', 'Цена за ед.', 'Общая сумма']

        values = []
        price_of_shiped_drugs = 0
        amount_of_shiped_drugs = Decimal('0')
        for index, shipment in enumerate(shipments): 
            value = [index+1, shipment.drug.name,
                     shipment.provider.name,
                     shipment.document,
                     shipment.producer.name,
                     shipment.serial_number,
                     shipment.date_of_comming.strftime("%d.%m.%Y"),
                     shipment.use_by_date.strftime("%d.%m.%Y"),
                     shipment.drug.unit.name,
                     shipment.initial_amount,
                     shipment.prise_for_unit,
                     shipment.total_price]
            amount_of_shiped_drugs += shipment.initial_amount
            price_of_shiped_drugs += shipment.total_price
            values.append(value)

        total_sum_row = {
            'I': 'Итого:',
            'J': amount_of_shiped_drugs,
            'L': price_of_shiped_drugs}
        columns_width = [3, 35, 20, 15, 21, 16, 11, 13, 12, 11, 11, 17]
        return Report([title], column_titles, [values], [total_sum_row], columns_width)

    @staticmethod
    def spent_drugs_report(from_date, to_date, department):
        spent_drugs_quersyet = Movement.objects.filter(date__gte=from_date, date__lte=to_date, department=department) \
                    .select_related('shipment', 'shipment__drug', 'shipment__drug__unit', 'department')

        title = ['Отчёт израсходованных лекарственных средств', f'с {from_date.strftime("%d.%m.%Y")} по {to_date.strftime("%d.%m.%Y")} года.', department.name]
        column_titles = ['№', 'Наименование', 'Фирма-поставщик', 'Фирма-изготовитель', 'Ед.измерения', 'Колич.', 'Цена за ед.', 'Общая сумма']
        columns_width = [3, 35, 20, 20, 12, 11, 11, 17]

        values = []
        price_of_spent_drugs = 0
        amount_of_spent_drugs = Decimal('0')

        for index, movement in enumerate(spent_drugs_quersyet): 
            value = [index+1,
                     movement.shipment.drug.name,
                     movement.shipment.provider.name,
                     movement.shipment.producer.name,
                     movement.shipment.drug.unit.name,
                     movement.amount,
                     movement.shipment.prise_for_unit,
                     movement.total_price]
            amount_of_spent_drugs += movement.amount
            price_of_spent_drugs += movement.total_price
            values.append(value)

        total_sum_row = {
            'E': 'Итого:',
            'F': amount_of_spent_drugs,
            'H': price_of_spent_drugs}
        return Report([title], column_titles, [values], [total_sum_row], columns_width)

    @staticmethod
    def spent_drugs_report_all_departments(from_date, to_date):
        spent_drugs_queryset = Movement.objects.filter(date__gte=from_date, date__lte=to_date) \
                    .select_related('shipment', 'shipment__drug', 'shipment__drug__unit', 'department') \
                    .order_by('department')

        title = [['Отчёт израсходованных лекарственных средств',
                  f'с {from_date.strftime("%d.%m.%Y")} по {to_date.strftime("%d.%m.%Y")} года для всех отделений.']]
        column_titles = ['№', 'Наименование', 'Фирма-поставщик', 'Фирма-изготовитель', 'Ед.измерения', 'Колич.',
                         'Цена за ед.', 'Общая сумма']
        columns_width = [3, 35, 20, 20, 12, 11, 11, 17]
        values = [[]]
        total_sum_row = [{}]
        if spent_drugs_queryset.exists():
            current_department = spent_drugs_queryset[0].department.name
            title[0].append(current_department)

            price_of_spent_drugs = 0
            amount_of_spent_drugs = Decimal('0')

            for index, movement in enumerate(spent_drugs_queryset):
                if current_department != movement.department.name:
                    current_department = movement.department.name
                    total_sum_row[len(values)-1] = {
                        'E': 'Итого:',
                        'F': amount_of_spent_drugs,
                        'H': price_of_spent_drugs}
                    values.append([])
                    total_sum_row.append([])
                    price_of_spent_drugs = 0
                    amount_of_spent_drugs = Decimal('0')
                    title.append([current_department])

                value = [index+1,
                         movement.shipment.drug.name,
                         movement.shipment.provider.name,
                         movement.shipment.producer.name,
                         movement.shipment.drug.unit.name,
                         movement.amount,
                         movement.shipment.prise_for_unit,
                         movement.total_price]
                amount_of_spent_drugs += movement.amount
                price_of_spent_drugs += movement.total_price

                values[len(values)-1].append(value)
            else:
                total_sum_row[len(values)-1] = {
                        'E': 'Итого:',
                        'F': amount_of_spent_drugs,
                        'H': price_of_spent_drugs}
        return Report(title, column_titles, values, total_sum_row, columns_width)

    @staticmethod
    def movement_report(from_date, to_date):
        class Drug:
            def __init__(self, drug_name, drug_unit,
                         amount_before=0, sum_before=0,
                         amount_shipped=0, sum_shipped=0,
                         amount_spent=0, sum_spent=0,
                         amount_write_off=0, sum_write_off=0):

                self.drug_name = drug_name
                self.drug_unit = drug_unit

                self.amount_before = amount_before
                self.sum_before = sum_before

                self.amount_shipped = amount_shipped
                self.sum_shipped = sum_shipped

                self.amount_spent = amount_spent
                self.sum_spent = sum_spent

                self.amount_write_off = amount_write_off
                self.sum_write_off = sum_write_off
            
            def get_data(self):
                self.final_amount = self.amount_before+self.amount_shipped-self.amount_spent-self.amount_write_off
                self.final_sum = self.sum_before+self.sum_shipped-self.sum_spent-self.sum_write_off
                return [
                    self.drug_name,
                    self.drug_unit,
                    self.amount_before,
                    self.sum_before,
                    self.amount_shipped,
                    self.sum_shipped,
                    self.amount_spent,
                    self.sum_spent,
                    self.amount_write_off,
                    self.sum_write_off,
                    self.final_amount,
                    self.final_sum,
                ]

        def namedtuplefetchall(cursor_data):
            """Return all rows from a cursor as a namedtuple"""
            nt_result = namedtuple('Result', [col[0] for col in cursor_data.description])
            return [nt_result(*row) for row in cursor_data.fetchall()]
        result = {}

        with connection.cursor() as cursor:
            # get remnants
            cursor.execute(
                """
                SELECT drug_id
                    ,drug_name
                    ,drug_unit
                    ,SUM(total_shipment_amount) OVER (PARTITION BY drug_id) as amount
                    ,SUM(total_shipment_price) OVER (PARTITION BY drug_id) as total_price
                FROM
                (
                    SELECT drugs_shipment.drug_id,
                        drugs_drug.NAME AS drug_name,
                        drugs_drugunit.NAME AS drug_unit,
                        drugs_shipment.initial_amount - SUM(IFNULL(drugs_movement.amount, "0")) OVER (partition BY drugs_shipment.id) AS total_shipment_amount,
                        (drugs_shipment.initial_amount - SUM(IFNULL(drugs_movement.amount, "0")) OVER (partition BY drugs_shipment.id)) * drugs_shipment.prise_for_unit AS total_shipment_price
                    FROM  drugs_shipment
                        JOIN drugs_drug
                            ON drugs_shipment.drug_id = drugs_drug.id
                        JOIN drugs_drugunit
                            ON drugs_drug.unit_id = drugs_drugunit.id
                        LEFT JOIN drugs_movement
                                ON ( drugs_movement.shipment_id = drugs_shipment.id
                                    AND drugs_movement.date < %s )
                    WHERE  drugs_shipment.date_of_comming < %s
                        AND ( drugs_shipment.date_of_run_out IS NULL
                                OR drugs_shipment.date_of_run_out > %s) 
                )
                GROUP BY drug_id
                """,
                [from_date, to_date, from_date],
            )
            remnants = namedtuplefetchall(cursor)
            
            # get movements
            cursor.execute(
                """
                SELECT drugs_shipment.initial_amount AS amount,
                    drugs_shipment.drug_id,
                    drugs_drug.name AS drug_name,
                    drugs_drugunit.name AS drug_unit,
                    SUM (drugs_movement.amount * drugs_shipment.prise_for_unit) AS total_price
                FROM drugs_shipment
                    JOIN drugs_drug ON drugs_shipment.drug_id = drugs_drug.id
                    JOIN drugs_drugunit ON drugs_drug.unit_id = drugs_drugunit.id
                    JOIN drugs_movement ON drugs_movement.shipment_id = drugs_shipment.id
                    WHERE drugs_shipment.date_of_comming >= %s AND
                        drugs_shipment.date_of_comming <= %s AND
                        (drugs_shipment.date_of_run_out IS NULL OR drugs_shipment.date_of_run_out >= %s)
                GROUP BY drugs_shipment.drug_id
                """,
                [from_date, to_date, from_date],
            )
            movements = namedtuplefetchall(cursor)

            # get write_offs
            cursor.execute(
                """
                SELECT drugs_shipment.initial_amount AS amount,
                    drugs_shipment.drug_id,
                    drugs_drug.name AS drug_name,
                    drugs_drugunit.name AS drug_unit,
                    SUM (drugs_writeoff.amount * drugs_shipment.prise_for_unit) AS total_price
                FROM drugs_shipment
                    JOIN drugs_drug ON drugs_shipment.drug_id = drugs_drug.id
                    JOIN drugs_drugunit ON drugs_drug.unit_id = drugs_drugunit.id
                    JOIN drugs_writeoff ON drugs_writeoff.shipment_id = drugs_shipment.id
                    WHERE drugs_shipment.date_of_comming >= %s AND
                        drugs_shipment.date_of_comming <= %s AND
                        (drugs_shipment.date_of_run_out IS NULL OR drugs_shipment.date_of_run_out >= %s)
                GROUP BY drugs_shipment.drug_id
                """,
                [from_date, to_date, from_date],
            )
            write_offs = namedtuplefetchall(cursor)

            # get shipments
            cursor.execute(
                """
                SELECT drugs_shipment.initial_amount AS amount,
                    drugs_shipment.drug_id,
                    drugs_drug.name AS drug_name,
                    drugs_drugunit.name AS drug_unit,
                    SUM (drugs_shipment.initial_amount * drugs_shipment.prise_for_unit) AS total_price
                FROM drugs_shipment
                    JOIN drugs_drug ON drugs_shipment.drug_id = drugs_drug.id
                    JOIN drugs_drugunit ON drugs_drug.unit_id = drugs_drugunit.id
                WHERE drugs_shipment.date_of_comming >= %s AND drugs_shipment.date_of_comming <= %s
                GROUP BY drugs_shipment.drug_id
                """,
                [from_date, to_date],
            )
            shipments = namedtuplefetchall(cursor)

            for remnant in remnants:
                result[remnant.drug_id] = Drug(remnant.drug_name, remnant.drug_unit,
                                               amount_before=remnant.amount,
                                               sum_before=remnant.total_price)

            for movement in movements:
                if movement.drug_id in result:
                    result[movement.drug_id].amount_spent = movement.amount
                    result[movement.drug_id].sum_spent = movement.total_price
                else:
                    result[movement.drug_id] = Drug(movement.drug_name, movement.drug_unit,
                                                    amount_spent=movement.amount,
                                                    sum_spent=movement.total_price)

            for write_off in write_offs:
                if write_off.drug_id in result:
                    result[write_off.drug_id].amount_write_off = write_off.amount
                    result[write_off.drug_id].sum_write_off = write_off.total_price
                else:
                    result[write_off.drug_id] = Drug(write_off.drug_name, write_off.drug_unit,
                                                     amount_write_off=write_off.amount,
                                                     sum_write_off=write_off.total_price)

            for shipment in shipments:
                if shipment.drug_id in result:
                    result[shipment.drug_id].amount_shipped = shipment.amount
                    result[shipment.drug_id].sum_shipped = shipment.total_price
                else:
                    result[shipment.drug_id] = Drug(shipment.drug_name, shipment.drug_unit,
                                                    amount_shipped=shipment.amount,
                                                    sum_shipped=shipment.total_price)
            '''for index in result:
                result[index].get_info()'''

        title = ['Отчёт о движении медикаментов', f'с {from_date.strftime("%d.%m.%Y")} по {to_date.strftime("%d.%m.%Y")} года.']
        column_titles = ['№', 'Наименование', 'Ед.измерения', 'Остаток на начало периода', '', 'Приход', '', 'Израсходовано', '', 'Списано', '','Остаток на конец периода', '']
        columns_width = [3, 35, 16, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12]

        values = [[]]

        shipment_sum = 0
        movement_sum = 0
        write_off_sum = 0
        remnants_sum = 0

        for index, drug_id in enumerate(result.keys()):

            value = [index+1,]+result[drug_id].get_data()

            shipment_sum += result[drug_id].sum_shipped
            movement_sum += result[drug_id].sum_spent
            write_off_sum += result[drug_id].sum_write_off
            remnants_sum += result[drug_id].final_sum
            values.append(value)

        total_sum_row = {
            'F': 'Итого:',
            'G': shipment_sum,
            'H': 'Итого:',
            'I': movement_sum,
            'J': 'Итого:',
            'K': write_off_sum,
            'L': 'Итого:',
            'M': remnants_sum,
        }

        return Report([title], column_titles, [values], [total_sum_row], columns_width, is_movement_report=True)

    def save(self, movement_report=True):
        work_book = Workbook()
        spreadsheet = work_book.active
        spreadsheet.title = "Отчет"
        bd = Side(border_style='thin')
        border = Border(left=bd, top=bd, right=bd, bottom=bd)
        last_column = get_column_letter(len(self.column_titles))
        current_row = 1
        for report_number in range(len(self.values)):
            
            #header
            for index, item in enumerate(self.title[report_number]):
                
                spreadsheet[f'A{index+current_row}'] = item
                spreadsheet[f'A{index+current_row}'].font = Font(bold=True, size=14)
                spreadsheet[f'A{index+current_row}'].alignment = Alignment(horizontal="center")
                spreadsheet.merge_cells(f'A{index+current_row}:{last_column}{index+current_row}')

            spreadsheet.append([])
            
            #column titles
            spreadsheet.append(self.column_titles)
            if self.is_movement_report:
                spreadsheet.append({
                    'D': 'Кол-во',
                    'E': 'Сумма',
                    'F': 'Кол-во',
                    'G': 'Сумма',
                    'H': 'Кол-во',
                    'I': 'Сумма',
                    'J': 'Кол-во',
                    'K': 'Сумма',
                    'L': 'Кол-во',
                    'M': 'Сумма',
                })
                spreadsheet.merge_cells('A4:A5')
                spreadsheet.merge_cells('B4:B5')
                spreadsheet.merge_cells('C4:C5')
                spreadsheet.merge_cells('D4:E4')
                spreadsheet.merge_cells('F4:G4')
                spreadsheet.merge_cells('H4:I4')
                spreadsheet.merge_cells('J4:K4')
                spreadsheet.merge_cells('L4:M4')
                
                for title_row_number in [4,5]:
                    for cell in spreadsheet[title_row_number]:
                        cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
                        cell.font = Font(bold=True, size=10)
                        cell.alignment = Alignment(horizontal="center", vertical="center")

            else:
                title_row_number = spreadsheet.max_row
                for cell in spreadsheet[title_row_number]:
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
                    cell.font = Font(bold=True, size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                spreadsheet.row_dimensions[title_row_number].height = 25


            
            # values
            for row in self.values[report_number]:
                row_number = spreadsheet.max_row + 1
                for column_index, value in enumerate(row):
                    column_letter = get_column_letter(column_index+1)
                    spreadsheet[f'{column_letter}{row_number}'] = value
                    spreadsheet[f'{column_letter}{row_number}'].border = border
                    spreadsheet[f'{column_letter}{row_number}'].font = Font(size=10)
            max_row = spreadsheet.max_row+1
            # total sum row
            
            for column in self.total_sum_row[report_number]:
                spreadsheet[f'{column}{max_row}'] = self.total_sum_row[report_number][column]
                spreadsheet[f'{column}{max_row}'].font = Font(bold=True, size=11)
                spreadsheet[f'{column}{max_row}'].border = border
            # savingsз file
            filename = 'report.xlsx'
            for index, width in enumerate(self.columns_width):
                spreadsheet.column_dimensions[get_column_letter(index+1)].width = width
            spreadsheet.append([])
            current_row = spreadsheet.max_row+1
        work_book.save(filename=filename)
